import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import ttk

class SpreadsheetGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Spreadsheet GUI")

        self.filename_label = tk.Label(master, text="No file selected")
        self.filename_label.pack()

        self.select_file_button = tk.Button(master, text="Select File", command=self.select_file)
        self.select_file_button.pack()

        self.input_label = tk.Label(master, text="Enter a value for Column A:")
        self.input_label.pack()

        self.input_entry = tk.Entry(master)
        self.input_entry.pack()
        self.input_entry.bind("<Return>", self.submit_value)

        self.tree = ttk.Treeview(master, columns=("A", "B", "C", "D"))
        self.tree.heading("#0", text="Index")
        self.tree.heading("A", text="Column A")
        self.tree.heading("B", text="Column B")
        self.tree.heading("C", text="Column C")
        self.tree.heading("D", text="Column D")
        self.tree.pack()

        self.tree.bind('<ButtonRelease-1>', self.select_cell)

    def select_file(self):
        file_path = filedialog.askopenfilename()
        self.filename_label.config(text=file_path)

        if file_path:
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            self.tree.delete(*self.tree.get_children())

            for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
                self.tree.insert("", "end", text=str(i), values=row)

    def submit_value(self, event):
        value = self.input_entry.get()
        self.input_entry.delete(0, tk.END)

        for item in self.tree.selection():
            self.tree.set(item, "D", "O")

    def select_cell(self, event):
        row_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if row_id and column == '#4':
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, self.tree.item(row_id)['values'][0])

root = tk.Tk()
spreadsheet_gui = SpreadsheetGUI(root)
root.mainloop()
