from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from datetime import datetime
from PyQt5 import QtCore
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QApplication,
    QLabel,
    QLineEdit,
    QPushButton,
    QSplitter,
    QTabWidget,
    QTreeWidget,
    QVBoxLayout,
    QWidget,
    QHBoxLayout,
    QTreeWidgetItem,
)
from PyQt5 import QtCore, QtGui
import PyPDF2
import win32print


class SpreadsheetGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Scan for SP IMPORT')
        self.previous_input = None
        self.pdf_path = None
        self.formatted_number = None

        # create tab widget
        self.tab_widget = QTabWidget()
        self.scan_tab = QWidget()
        self.inspection_tab = QWidget()
        self.sagawa_tab = QWidget()
        self.tab_widget.addTab(self.scan_tab, "SCAN")
        self.tab_widget.addTab(self.inspection_tab, "INSPECTION")
        self.tab_widget.addTab(self.sagawa_tab, "SAGAWA")

        # create widgets for scan tab
        self.filename_label = QLabel('No file selected')
        self.select_file_button = QPushButton('Select File (F2)')
        self.select_file_button.clicked.connect(self.select_file)

        self.input_label = QLabel('HOUSE NO.를 스캔하세요.')
        self.input_entry = QLineEdit()
        self.input_entry.returnPressed.connect(self.submit_value)

        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(['NO.', 'HOUSE NO.', 'CT', 'Scanned', 'COMPLETE', 'TIME', 'shipper', '区分'])
        self.tree.setColumnCount(8)

        #set column widths
        self.tree.setColumnWidth(0, 50)
        self.tree.setColumnWidth(1, 120)
        self.tree.setColumnWidth(2, 70)
        self.tree.setColumnWidth(3, 70)
        self.tree.setColumnWidth(4, 70)
        self.tree.setColumnWidth(5, 120)
        # self.tree.setColumnWidth(6, 120)
        self.tree.setColumnWidth(7, 50)

        # set stylesheet to show grid lines
        self.tree.setStyleSheet("QTreeView::item { border-bottom: 1px solid black; border-right: 1px solid black; }"
                        "QTreeView::item:selected { background-color: #d8d8d8; }")

        # enable sorting
        self.tree.setSortingEnabled(True)
        self.tree.sortItems(0, QtCore.Qt.AscendingOrder)

        scan_layout = QHBoxLayout()
        splitter = QSplitter(QtCore.Qt.Horizontal)

        left_widget = QWidget()
        left_layout = QVBoxLayout()
        left_layout.addWidget(self.filename_label)
        left_layout.addWidget(self.select_file_button)
        left_layout.addWidget(self.input_label)
        left_layout.addWidget(self.input_entry)
        left_layout.addWidget(self.tree)

        # create a save button and add it to the layout
        self.save_button = QPushButton('Save (F5)')
        self.save_button.clicked.connect(self.save_data)
        left_layout.addWidget(self.save_button)

        left_widget.setLayout(left_layout)

        right_widget = QWidget()
        right_layout = QVBoxLayout()

        # create a label to show the selected cell value
        self.cell_value_label = QLabel()
        self.cell_value_label.setAlignment(Qt.AlignCenter)  #오른쪽 위젯의 정중앙에 값 출력
        right_layout.addWidget(self.cell_value_label)

        right_layout.addWidget(QLabel(""))
        
        self.cell_value_label2 = QLabel()
        self.cell_value_label2.setAlignment(Qt.AlignCenter)
        right_layout.addWidget(self.cell_value_label2)

        self.back_button = QPushButton('Back (F4)')
        self.back_button.clicked.connect(self.back_one_step)
        left_layout.addWidget(self.back_button)

        right_widget.setLayout(right_layout)

        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)

        scan_layout.addWidget(splitter)
        self.scan_tab.setLayout(scan_layout)


        # create widgets for inspection tab
        self.inspection_filename_label = QLabel('No file selected')
        self.inspection_select_file_button = QPushButton('Select File (F2)')
        self.inspection_select_file_button.clicked.connect(self.select_file2)

        self.inspection_input_label = QLabel('HOUSE NO.를 스캔하세요.')
        self.inspection_input_entry = QLineEdit()
        self.inspection_input_entry.returnPressed.connect(self.submit_value2)

        self.inspection_tree = QTreeWidget()
        self.inspection_tree.setHeaderLabels(['NO.', 'HOUSE NO.', '出力', '開被', 'TIME'])
        self.inspection_tree.setColumnCount(5)

        # set column widths
        self.inspection_tree.setColumnWidth(0, 50)
        self.inspection_tree.setColumnWidth(1, 120)
        self.inspection_tree.setColumnWidth(2, 70)
        self.inspection_tree.setColumnWidth(3, 70)
        self.inspection_tree.setColumnWidth(4, 70)
 

        # set stylesheet to show grid lines
        self.inspection_tree.setStyleSheet("QTreeView::item { border-bottom: 1px solid black; border-right: 1px solid black; }"
                                "QTreeView::item:selected { background-color: #d8d8d8; }")

        # enable sorting
        self.inspection_tree.setSortingEnabled(True)
        self.inspection_tree.sortItems(0, QtCore.Qt.AscendingOrder)

        inspection_layout = QHBoxLayout()
        splitter = QSplitter(QtCore.Qt.Horizontal)

        inspection_left_widget = QWidget()
        inspection_left_layout = QVBoxLayout()
        inspection_left_layout.addWidget(self.inspection_filename_label)
        inspection_left_layout.addWidget(self.inspection_select_file_button)
        inspection_left_layout.addWidget(self.inspection_input_label)
        inspection_left_layout.addWidget(self.inspection_input_entry)
        inspection_left_layout.addWidget(self.inspection_tree)

        # create a save button and add it to the layout
        self.inspection_save_button = QPushButton('Save (F5)')
        self.inspection_save_button.clicked.connect(self.save_data2)
        inspection_left_layout.addWidget(self.inspection_save_button)

        inspection_left_widget.setLayout(inspection_left_layout)

        inspection_right_widget = QWidget()
        inspection_right_layout = QVBoxLayout()

        # create a label to show the selected cell value
        self.inspection_cell_value_label = QLabel()
        self.inspection_cell_value_label.setAlignment(Qt.AlignCenter)  #오른쪽 위젯의 정중앙에 값 출력
        inspection_right_layout.addWidget(self.inspection_cell_value_label)

        inspection_right_layout.addWidget(QLabel(""))
                
        self.inspection_cell_value_label2 = QLabel()
        self.inspection_cell_value_label2.setAlignment(Qt.AlignCenter)
        inspection_right_layout.addWidget(self.inspection_cell_value_label2)

        self.inspection_back_button = QPushButton('Back (F4)')
        self.inspection_back_button.clicked.connect(self.back_one_step)
        inspection_left_layout.addWidget(self.inspection_back_button)

        inspection_right_widget.setLayout(inspection_right_layout)

        splitter.addWidget(inspection_left_widget)
        splitter.addWidget(inspection_right_widget)

        inspection_layout.addWidget(splitter)
        self.inspection_tab.setLayout(inspection_layout)


        # create widgets for sagawa tab
        sagawa_layout = QVBoxLayout()
        sagawa_layout.addWidget(QLabel("SAGAWA TAB"))

        self.sagawa_tab.setLayout(sagawa_layout)

        layout = QVBoxLayout()
        layout.addWidget(self.tab_widget)

        self.setLayout(layout)

        self.setGeometry(100, 100, 1200, 600)

    def back_one_step(self):
        # get the previous input value
        previous_input = self.previous_input

        # find the item with the previous input value in the tree
        items = self.tree.findItems(previous_input, Qt.MatchExactly, 1)

        if len(items) > 0:
            item = items[-1]  # get the last matching item
            scanned_col = int(item.text(3))

            # decrement scanned column by 1
            if scanned_col > 1:
                scanned_col -= 1
                item.setText(3, str(scanned_col))

                # revert completion column to blank
                item.setText(4, '')

                # # clear timestamp column
                # item.setText(5, '')

            # remove scanned_col if it is 1
            elif scanned_col == 1:
                item.setText(3, '')
                item.setText(4, '')
                item.setText(5, '')

                # clear 구분값 and 쉬퍼값 labels
                self.cell_value_label.setText('')
                self.cell_value_label2.setText('')




    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_F5:
            self.save_data()
        if event.key() == QtCore.Qt.Key_F2:
            self.select_file()
        if event.key() == QtCore.Qt.Key_F4:
            self.back_one_step()

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select File')
        if file_path:
            self.filename_label.setText(file_path)
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            self.tree.clear()

            for i, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True), start=1):
                item = QTreeWidgetItem([str(i), str(row[0]), str(row[1]), '', ''])
                item.setText(1, str(row[0]))
                item.setText(2, str(row[1]))
                item.setText(6, str(row[2]))
                item.setText(7, str(row[3]))
                self.tree.addTopLevelItem(item)

            # set focus on the input_entry
            self.input_entry.setFocus()



    def submit_value(self):
        value = self.input_entry.text()

        # get the current value of input_entry
        current_input = self.input_entry.text()

        # add the current value to the tree
        # self.add_to_tree(current_input)

        # save the current value for the back button
        self.previous_input = current_input
        self.input_entry.clear()

        is_value_exists = False

        for item in self.tree.findItems(value, Qt.MatchExactly, 1):
            # increment scanned column by 1
            scanned_col = item.text(3)
            if scanned_col:
                scanned_col = int(scanned_col) + 1
            else:
                scanned_col = 1
            item.setText(3, str(scanned_col))

            # set 'O' to complete column if Ct and scanned values are equal
            ct_col = item.text(2)
            if ct_col and str(scanned_col) == ct_col:
                item.setText(4, 'O')
                item.setForeground(4, QtGui.QColor('red'))
                font = QtGui.QFont()
                font.setBold(True)
                item.setFont(4, font)
            elif scanned_col > int(ct_col):
                QMessageBox.warning(self, 'Error', '해당 HOUSE NO.의 총 CT수를 초과했습니다.')
                item.setText(3, str(int(ct_col)))  # revert scanned value to Ct value

            # set current timestamp to column C
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            item.setText(5, timestamp)

            is_value_exists = True

        if not is_value_exists:
            QApplication.beep()  # 비프음 재생
            QMessageBox.warning(self, 'Error', '없는 HOUSE NO.입니다.')
        
        if is_value_exists:
            # Scroll to the last scanned item
            last_scanned_item = self.tree.findItems(value, Qt.MatchExactly, 1)[-1]
            self.tree.scrollToItem(last_scanned_item)

            #구분값 출력
            font = QtGui.QFont()
            font.setPointSize(400)
            self.cell_value_label.setFont(font)
            self.cell_value_label.setText(f"{item.text(7)}")

            #쉬퍼값 출력
            font = QtGui.QFont()
            font.setPointSize(100)
            self.cell_value_label2.setFont(font)
            self.cell_value_label2.setText(f"{item.text(6)}")

    def save_data(self):
        # get file path and name
        file_path = self.filename_label.text()
        if not file_path:
            QMessageBox.warning(self, 'Error', 'Please select a file first.')
            return
        file_dir, file_name = os.path.split(file_path)
        file_name_no_ext, file_ext = os.path.splitext(file_name)

        # create new workbook and worksheet
        new_file_name = file_name_no_ext + '_new' + file_ext
        new_file_path, _ = QFileDialog.getSaveFileName(self, 'Save File As', new_file_name, 'Excel files (*.xlsx)')
        if not new_file_path:
            return
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Data'

        # write column headers
        headers = ['NO.', 'HOUSE NO.', 'CT', 'Scanned', 'COMPLETE', 'TIME']
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header

        # write data to worksheet
        for row_num in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(row_num)
            house_no_value = item.text(0)
            name_value = item.text(1)
            ct_value = item.text(2)
            scan_value = item.text(3)
            complete_value = item.text(4)
            time_value = item.text(5)

            # write values to worksheet
            row = [house_no_value, name_value, ct_value, scan_value, complete_value, time_value]
            worksheet.append(row)

        # save workbook to new file path
        workbook.save(new_file_path)
        QMessageBox.information(self, 'Saved', 'Data has been saved to {}.'.format(new_file_path))

##################################################
    def select_file2(self):
        file_dialog = QFileDialog()
        file_dialog.setDefaultSuffix("pdf")
        file_dialog.setNameFilters(["PDF Files (*.pdf)"])
        file_path, _ = file_dialog.getOpenFileName(parent=self, caption='Open file', directory=os.getcwd(), filter='PDF Files (*.pdf)')

        if file_path:
            self.inspection_filename_label.setText(file_path)
            self.pdf_path = file_path
        
        self.inspection_input_entry.setFocus()


    def find_pages_with_number(self, pdf_path, number):
        pdf_file = open(pdf_path, 'rb')
        pdf_reader = PyPDF2.PdfFileReader(pdf_file)

        pages_with_number = []

        for page_num in range(pdf_reader.getNumPages()):
            page = pdf_reader.getPage(page_num)
            page_text = page.extractText()
            if number in page_text:
                pages_with_number.append(page_num)

        pdf_file.close()

        return pages_with_number

    def create_printable_pdf(self, pdf_path, pages_to_print):
        output_folder = "output"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        output_pdf_path = os.path.join(output_folder, "printable.pdf")
        output_pdf = PyPDF2.PdfFileWriter()

        pdf_file = open(pdf_path, 'rb')
        pdf_reader = PyPDF2.PdfFileReader(pdf_file)

        for page_num in pages_to_print:
            page = pdf_reader.getPage(page_num)
            output_pdf.addPage(page)

        output_pdf_stream = open(output_pdf_path, 'wb')
        output_pdf.write(output_pdf_stream)
        output_pdf_stream.close()

        pdf_file.close()

        return output_pdf_path


    def print_pages_with_number(self, formatted_number):
        if not self.pdf_path:
            QMessageBox.warning(self, "Error", "PDF 파일을 먼저 선택해주세요.")
            return
        pages_with_number = self.find_pages_with_number(self.pdf_path, formatted_number)
        printer_name = win32print.GetDefaultPrinter()

        if not pages_with_number:
            QMessageBox.warning(self, "Error", f"HOUSE NO.{formatted_number}가 존재하지 않습니다.")
            return

        # 이미 출력된 HOUSE NO.인지 확인
        for i in range(self.inspection_tree.topLevelItemCount()):
            if formatted_number == self.inspection_tree.topLevelItem(i).text(1):
                QMessageBox.warning(self, "Error", f"HOUSE NO.{formatted_number}는 이미 출력되었습니다.")
                return

        h_printer = win32print.OpenPrinter(printer_name)
        doc_info = ("Pages with Number", None, "RAW")

        output_pdf_path = self.create_printable_pdf(self.pdf_path, pages_with_number)

        with open(output_pdf_path, "rb") as pdf_file:
            pdf_data = pdf_file.read()

            job_id = win32print.StartDocPrinter(h_printer, 1, doc_info)
            win32print.StartPagePrinter(h_printer)
            win32print.WritePrinter(h_printer, pdf_data)
            win32print.EndPagePrinter(h_printer)
            win32print.EndDocPrinter(h_printer)

        win32print.ClosePrinter(h_printer)

        # 출력한 정보를 TreeWidget에 추가
        row = QTreeWidgetItem([str(self.inspection_tree.topLevelItemCount()+1), formatted_number, "O", "", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        self.inspection_tree.addTopLevelItem(row)
        

    def submit_value2(self):
        input_value = self.inspection_input_entry.text()
        # self.formatted_number = "KIT2" + input_value.strip()
        self.formatted_number = input_value.strip()
        self.inspection_input_entry.clear()
        
        self.print_pages_with_number(self.formatted_number)

    def save_data2(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Save Excel file", "", "Excel files (*.xlsx)")
        if not filename:
            return

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        
        # write header row
        header_labels = [self.inspection_tree.headerItem().text(i) for i in range(self.inspection_tree.columnCount())]
        for i, label in enumerate(header_labels):
            worksheet.cell(row=1, column=i+1, value=label)

        # write data rows
        for i in range(self.inspection_tree.topLevelItemCount()):
            row_items = [self.inspection_tree.topLevelItem(i).text(j) for j in range(self.inspection_tree.columnCount())]
            for j, item in enumerate(row_items):
                worksheet.cell(row=i+2, column=j+1, value=item)

        try:
            workbook.save(filename)
            QMessageBox.information(self, "Save Successful", "Data saved successfully.")
        except Exception as e:
            QMessageBox.warning(self, "Save Failed", f"Data save failed. Error: {e}")
        finally:
            workbook.close()


if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    spreadsheet_gui = SpreadsheetGUI()
    spreadsheet_gui.show()
    sys.exit(app.exec_())


# pyinstaller -w -F --add-data '*.png;.' 5_add...............