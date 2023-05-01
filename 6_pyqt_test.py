from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from datetime import datetime
from PyQt5 import QtCore
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class SpreadsheetGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Spreadsheet GUI')

        self.filename_label = QLabel('No file selected')
        self.select_file_button = QPushButton('Select File')
        self.select_file_button.clicked.connect(self.select_file)

        self.input_label = QLabel('Enter a value for Column A:')
        self.input_entry = QLineEdit()
        self.input_entry.returnPressed.connect(self.submit_value)

        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(['NO.', 'HOUSE NO.', 'SCAN', 'TIME'])
        self.tree.setColumnCount(4)


        # set column widths
        self.tree.setColumnWidth(0, 50)
        self.tree.setColumnWidth(1, 150)
        self.tree.setColumnWidth(2, 150)
        self.tree.setColumnWidth(3, 150)

        # set stylesheet to show grid lines
        self.tree.setStyleSheet("QTreeView::item { border-bottom: 1px solid black; border-right: 1px solid black; }"
                        "QTreeView::item:selected { background-color: #d8d8d8; }")

        # enable sorting
        self.tree.setSortingEnabled(True)
        self.tree.sortItems(0, QtCore.Qt.AscendingOrder)

        layout = QVBoxLayout()
        layout.addWidget(self.filename_label)
        layout.addWidget(self.select_file_button)
        layout.addWidget(self.input_label)
        layout.addWidget(self.input_entry)
        layout.addWidget(self.tree)

        self.setLayout(layout)

        self.tree.itemClicked.connect(self.select_cell)


        # create a save button and add it to the layout
        self.save_button = QPushButton('Save')
        self.save_button.clicked.connect(self.save_data)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        self.tree.itemClicked.connect(self.select_cell)        

        # set GUI size
        self.setGeometry(100, 100, 800, 600)

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select File')
        if file_path:
            self.filename_label.setText(file_path)
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            self.tree.clear()

            for i, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=1):
                item = QTreeWidgetItem([str(i), str(row[0]), '', ''])
                self.tree.addTopLevelItem(item)
                

    def submit_value(self):
        value = self.input_entry.text()
        self.input_entry.clear()

        is_value_exists = False

        for item in self.tree.findItems(value, Qt.MatchExactly, 1):
            # set 'O' to column B
            item.setText(2, 'O')
            # set current timestamp to column C
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            item.setText(3, timestamp)
            is_value_exists = True

        if not is_value_exists:
            QApplication.beep()  # 비프음 재생
            QMessageBox.warning(self, 'Error', 'Value not found in Column A.')


    def select_cell(self, item):
        if self.tree.currentColumn() == 3:
            self.input_entry.setText(item.text(1))

    def save_data(self):
        # get file path and name
        file_path = self.filename_label.text()
        if not file_path:
            QMessageBox.warning(self, 'Error', 'Please select a file first.')
            return
        file_dir, file_name = os.path.split(file_path)
        file_name_no_ext, file_ext = os.path.splitext(file_name)

        # create new workbook and worksheet
        new_file_name = file_name_no_ext + '_new' + file_ext
        new_file_path = os.path.join(file_dir, new_file_name)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Data'

        # write column headers
        headers = ['HOUSE NO.', 'NAME', 'SCAN', 'TIME']
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header

        # write data to worksheet
        for row_num in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(row_num)
            house_no_value = item.text(0)
            name_value = item.text(1)
            scan_value = item.text(2)
            time_value = item.text(3)

            row_values = [house_no_value, name_value, scan_value, time_value]
            for col_num, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_num+1, column=col_num)
                cell.value = value

        # save workbook
        workbook.save(new_file_path)
        QMessageBox.information(self, 'Save', 'Data saved successfully.')

if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    spreadsheet_gui = SpreadsheetGUI()
    spreadsheet_gui.show()
    sys.exit(app.exec_())

# 작업중 표를 new 붙여서 자동 저장하는 버전
