from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from datetime import datetime
from PyQt5 import QtCore
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class SpreadsheetGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Scan for SP IMPORT')

        # create tab widget
        self.tab_widget = QTabWidget()
        self.scan_tab = QWidget()
        self.inspection_tab = QWidget()
        self.sagawa_tab = QWidget()
        self.tab_widget.addTab(self.scan_tab, "SCAN")
        self.tab_widget.addTab(self.inspection_tab, "INSPECTION")
        self.tab_widget.addTab(self.sagawa_tab, "SAGAWA")

        # create widgets for scan tab
        self.filename_label = QLabel('No file selected')
        self.select_file_button = QPushButton('Select File (F2)')
        self.select_file_button.clicked.connect(self.select_file)

        self.input_label = QLabel('HOUSE NO.를 스캔하세요.')
        self.input_entry = QLineEdit()
        self.input_entry.returnPressed.connect(self.submit_value)

        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(['NO.', 'HOUSE NO.', 'CT', 'Scanned','COMPLETE', 'TIME'])
        self.tree.setColumnCount(6)

        # set column widths
        self.tree.setColumnWidth(0, 50)
        self.tree.setColumnWidth(1, 150)
        self.tree.setColumnWidth(2, 150)
        self.tree.setColumnWidth(3, 150)

        # set stylesheet to show grid lines
        self.tree.setStyleSheet("QTreeView::item { border-bottom: 1px solid black; border-right: 1px solid black; }"
                        "QTreeView::item:selected { background-color: #d8d8d8; }")

        # enable sorting
        self.tree.setSortingEnabled(True)
        self.tree.sortItems(0, QtCore.Qt.AscendingOrder)

        scan_layout = QVBoxLayout()
        scan_layout.addWidget(self.filename_label)
        scan_layout.addWidget(self.select_file_button)
        scan_layout.addWidget(self.input_label)
        scan_layout.addWidget(self.input_entry)
        scan_layout.addWidget(self.tree)

        # create a save button and add it to the layout
        self.save_button = QPushButton('Save (F5)')
        self.save_button.clicked.connect(self.save_data)
        scan_layout.addWidget(self.save_button)

        self.scan_tab.setLayout(scan_layout)

        # create widgets for inspection tab
        inspection_layout = QVBoxLayout()
        inspection_layout.addWidget(QLabel("INSPECTION TAB"))

        self.inspection_tab.setLayout(inspection_layout)

        # create widgets for sagawa tab
        sagawa_layout = QVBoxLayout()
        sagawa_layout.addWidget(QLabel("SAGAWA TAB"))

        self.sagawa_tab.setLayout(sagawa_layout)

        layout = QVBoxLayout()
        layout.addWidget(self.tab_widget)

        self.setLayout(layout)

        self.tree.itemClicked.connect(self.select_cell)
        self.setGeometry(100, 100, 800, 600)


    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_F5:
            self.save_data()
        if event.key() == QtCore.Qt.Key_F2:
            self.select_file()

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select File')
        if file_path:
            self.filename_label.setText(file_path)
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            self.tree.clear()

            for i, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True), start=1):
                item = QTreeWidgetItem([str(i), str(row[0]), str(row[1]), '', ''])
                self.tree.addTopLevelItem(item)

            # set focus on the input_entry
            self.input_entry.setFocus()
         

    def submit_value(self):
        value = self.input_entry.text()
        self.input_entry.clear()

        is_value_exists = False

        for item in self.tree.findItems(value, Qt.MatchExactly, 1):
            # increment scanned column by 1
            scanned_col = item.text(3)
            if scanned_col:
                scanned_col = int(scanned_col) + 1
            else:
                scanned_col = 1
            item.setText(3, str(scanned_col))

            # set 'O' to complete column if Ct and scanned values are equal
            ct_col = item.text(2)
            if ct_col and str(scanned_col) == ct_col:
                item.setText(4, 'O')
            elif scanned_col > int(ct_col):
                QMessageBox.warning(self, 'Error', '해당 HOUSE NO.의 총 CT수를 초과했습니다.')
                item.setText(3, str(int(ct_col)))  # revert scanned value to Ct value

            # set current timestamp to column C
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            item.setText(5, timestamp)

            is_value_exists = True

        if not is_value_exists:
            QApplication.beep()  # 비프음 재생
            QMessageBox.warning(self, 'Error', '없는 HOUSE NO.입니다.')
        
        if is_value_exists:
            # Scroll to the last scanned item
            last_scanned_item = self.tree.findItems(value, Qt.MatchExactly, 1)[-1]
            self.tree.scrollToItem(last_scanned_item)


    def select_cell(self, item):
        if self.tree.currentColumn() == 3:
            self.input_entry.setText(item.text(1))

    def save_data(self):
        # get file path and name
        file_path = self.filename_label.text()
        if not file_path:
            QMessageBox.warning(self, 'Error', 'Please select a file first.')
            return
        file_dir, file_name = os.path.split(file_path)
        file_name_no_ext, file_ext = os.path.splitext(file_name)

        # create new workbook and worksheet
        new_file_name = file_name_no_ext + '_new' + file_ext
        new_file_path, _ = QFileDialog.getSaveFileName(self, 'Save File As', new_file_name, 'Excel files (*.xlsx)')
        if not new_file_path:
            return
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Data'

        # write column headers
        headers = ['NO.', 'HOUSE NO.', 'CT', 'Scanned', 'COMPLETE', 'TIME']
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header

        # write data to worksheet
        for row_num in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(row_num)
            house_no_value = item.text(0)
            name_value = item.text(1)
            scan_value = item.text(3)
            complete_value = item.text(4)
            time_value = item.text(5)

            # write values to worksheet
            row = [house_no_value, name_value, '', scan_value, complete_value, time_value]
            worksheet.append(row)

        # save workbook to new file path
        workbook.save(new_file_path)
        QMessageBox.information(self, 'Saved', 'Data has been saved to {}.'.format(new_file_path))


if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    spreadsheet_gui = SpreadsheetGUI()
    spreadsheet_gui.show()
    sys.exit(app.exec_())


# pyinstaller -w -F --add-data '*.png;.' 5_add...............